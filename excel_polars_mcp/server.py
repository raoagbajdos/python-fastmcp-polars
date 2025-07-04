"""MCP Server for converting Excel files to Polars DataFrames."""

import asyncio
import json
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

import polars as pl
from fastmcp import FastMCP
from pydantic import BaseModel


class ReadExcelArgs(BaseModel):
    """Arguments for reading Excel file."""
    file_path: str
    sheet_name: Optional[str] = None
    has_header: bool = True
    infer_schema_length: int = 100


class ListSheetsArgs(BaseModel):
    """Arguments for listing Excel sheets."""
    file_path: str


class ReadExcelSheetArgs(BaseModel):
    """Arguments for reading specific Excel sheet."""
    file_path: str
    sheet_name: str
    has_header: bool = True
    infer_schema_length: int = 100


# Create FastMCP server
mcp = FastMCP("Excel to Polars Converter")


@mcp.tool()
async def read_excel(args: ReadExcelArgs) -> Dict[str, Any]:
    """
    Read an Excel file and convert it to Polars DataFrame format.
    
    Args:
        args: ReadExcelArgs containing file_path, optional sheet_name, 
              has_header flag, and infer_schema_length
    
    Returns:
        Dictionary containing the DataFrame data and metadata
    """
    try:
        file_path = Path(args.file_path)
        
        if not file_path.exists():
            return {"error": f"File not found: {args.file_path}"}
        
        if not file_path.suffix.lower() in ['.xlsx', '.xls']:
            return {"error": "File must be an Excel file (.xlsx or .xls)"}
        
        # Read Excel file with Polars
        df = pl.read_excel(
            source=args.file_path,
            sheet_name=args.sheet_name,
            has_header=args.has_header,
            infer_schema_length=args.infer_schema_length
        )
        
        # Convert to dictionary format
        result = {
            "success": True,
            "data": df.to_dict(as_series=False),
            "schema": {col: str(dtype) for col, dtype in df.schema.items()},
            "shape": df.shape,
            "sheet_name": args.sheet_name,
            "columns": df.columns
        }
        
        return result
        
    except Exception as e:
        return {"error": f"Failed to read Excel file: {str(e)}"}


@mcp.tool()
async def list_sheets(args: ListSheetsArgs) -> Dict[str, Any]:
    """
    List all sheet names in an Excel file.
    
    Args:
        args: ListSheetsArgs containing file_path
    
    Returns:
        Dictionary containing list of sheet names
    """
    try:
        file_path = Path(args.file_path)
        
        if not file_path.exists():
            return {"error": f"File not found: {args.file_path}"}
        
        if not file_path.suffix.lower() in ['.xlsx', '.xls']:
            return {"error": "File must be an Excel file (.xlsx or .xls)"}
        
        # Get sheet names using openpyxl for .xlsx files
        if file_path.suffix.lower() == '.xlsx':
            from openpyxl import load_workbook
            workbook = load_workbook(filename=args.file_path, read_only=True)
            sheets = workbook.sheetnames
            workbook.close()
        else:
            # For .xls files, we'll try to read and catch the error to get sheet info
            try:
                # This is a workaround as polars doesn't directly provide sheet listing
                pl.read_excel(source=args.file_path, sheet_name=0)
                sheets = ["Sheet1"]  # Default assumption for .xls
            except Exception:
                sheets = []
        
        return {
            "success": True,
            "sheets": sheets,
            "file_path": args.file_path
        }
        
    except Exception as e:
        return {"error": f"Failed to list sheets: {str(e)}"}


@mcp.tool()
async def read_excel_sheet(args: ReadExcelSheetArgs) -> Dict[str, Any]:
    """
    Read a specific sheet from an Excel file and convert to Polars DataFrame.
    
    Args:
        args: ReadExcelSheetArgs containing file_path, sheet_name, 
              has_header flag, and infer_schema_length
    
    Returns:
        Dictionary containing the DataFrame data and metadata for the specific sheet
    """
    try:
        file_path = Path(args.file_path)
        
        if not file_path.exists():
            return {"error": f"File not found: {args.file_path}"}
        
        if not file_path.suffix.lower() in ['.xlsx', '.xls']:
            return {"error": "File must be an Excel file (.xlsx or .xls)"}
        
        # Read specific sheet with Polars
        df = pl.read_excel(
            source=args.file_path,
            sheet_name=args.sheet_name,
            has_header=args.has_header,
            infer_schema_length=args.infer_schema_length
        )
        
        # Convert to dictionary format
        result = {
            "success": True,
            "data": df.to_dict(as_series=False),
            "schema": {col: str(dtype) for col, dtype in df.schema.items()},
            "shape": df.shape,
            "sheet_name": args.sheet_name,
            "columns": df.columns
        }
        
        return result
        
    except Exception as e:
        return {"error": f"Failed to read Excel sheet '{args.sheet_name}': {str(e)}"}


def main() -> None:
    """Run the MCP server."""
    mcp.run()


if __name__ == "__main__":
    main()