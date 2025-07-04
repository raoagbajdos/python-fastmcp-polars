"""Convert actuarial Excel file to Polars format using direct Polars methods."""

import json
from pathlib import Path

import polars as pl
from openpyxl import load_workbook


def convert_excel_to_polars(excel_path: str, output_dir: str):
    """Convert Excel file to Polars format and save outputs."""
    excel_file = Path(excel_path)
    output_path = Path(output_dir)
    
    if not excel_file.exists():
        print(f"❌ Excel file not found: {excel_path}")
        return
    
    output_path.mkdir(exist_ok=True)
    
    print(f"📖 Reading Excel file: {excel_file}")
    
    # Get list of sheets using openpyxl
    try:
        workbook = load_workbook(filename=str(excel_file), read_only=True)
        sheets = workbook.sheetnames
        workbook.close()
    except Exception as e:
        print(f"❌ Failed to read Excel file: {e}")
        return
    
    print(f"📋 Found {len(sheets)} sheets: {', '.join(sheets)}")
    
    # Process each sheet
    for sheet_name in sheets:
        print(f"\n🔄 Processing sheet: {sheet_name}")
        
        try:
            # Read sheet data using polars
            df = pl.read_excel(
                source=str(excel_file),
                sheet_name=sheet_name,
                has_header=True
            )
            
            print(f"   📊 Shape: {df.shape}")
            print(f"   📝 Columns: {', '.join(df.columns)}")
            
            # Save in multiple formats
            base_name = sheet_name.lower()
            
            # 1. Parquet (efficient binary format)
            parquet_path = output_path / f"{base_name}.parquet"
            df.write_parquet(parquet_path)
            print(f"   💾 Saved Parquet: {parquet_path}")
            
            # 2. CSV (human readable)
            csv_path = output_path / f"{base_name}.csv"
            df.write_csv(csv_path)
            print(f"   💾 Saved CSV: {csv_path}")
            
            # 3. JSON (structured data)
            json_path = output_path / f"{base_name}.json"
            df.write_json(json_path)
            print(f"   💾 Saved JSON: {json_path}")
            
            # 4. Schema information
            schema_path = output_path / f"{base_name}_schema.json"
            schema_info = {
                "sheet_name": sheet_name,
                "shape": df.shape,
                "columns": df.columns,
                "dtypes": {col: str(dtype) for col, dtype in zip(df.columns, df.dtypes)},
                "sample_data": df.head(3).to_dicts(),
                "statistics": get_basic_stats(df)
            }
            
            with open(schema_path, 'w') as f:
                json.dump(schema_info, f, indent=2, default=str)
            print(f"   📋 Saved Schema: {schema_path}")
            
        except Exception as e:
            print(f"   ❌ Error processing sheet {sheet_name}: {e}")
    
    print(f"\n✅ Conversion completed! Output saved to: {output_path}")
    
    # Create summary report
    create_summary_report(output_path, sheets)


def get_basic_stats(df: pl.DataFrame) -> dict:
    """Get basic statistics for the DataFrame."""
    stats = {}
    
    for col in df.columns:
        col_stats = {"type": str(df[col].dtype)}
        
        if df[col].dtype in [pl.Int64, pl.Int32, pl.Float64, pl.Float32]:
            try:
                col_stats.update({
                    "min": df[col].min(),
                    "max": df[col].max(),
                    "mean": df[col].mean(),
                    "null_count": df[col].null_count()
                })
            except Exception:
                pass
        elif df[col].dtype == pl.String:
            col_stats.update({
                "unique_count": df[col].n_unique(),
                "null_count": df[col].null_count()
            })
        
        stats[col] = col_stats
    
    return stats


def create_summary_report(output_path: Path, sheets: list):
    """Create a summary report of the conversion."""
    from datetime import datetime
    
    report_path = output_path / "conversion_summary.md"
    
    with open(report_path, 'w') as f:
        f.write("# Actuarial Data Conversion Summary\n\n")
        f.write(f"**Conversion Date:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
        f.write("**Source:** actuarial_data.xlsx\n")
        f.write(f"**Output Directory:** {output_path}\n\n")
        
        f.write("## Sheets Processed\n\n")
        for sheet in sheets:
            base_name = sheet.lower()
            f.write(f"### {sheet}\n")
            f.write(f"- **Parquet:** `{base_name}.parquet`\n")
            f.write(f"- **CSV:** `{base_name}.csv`\n")
            f.write(f"- **JSON:** `{base_name}.json`\n")
            f.write(f"- **Schema:** `{base_name}_schema.json`\n\n")
        
        f.write("## File Formats\n\n")
        f.write("- **Parquet:** Efficient binary format, best for data analysis\n")
        f.write("- **CSV:** Human-readable, compatible with Excel and other tools\n")
        f.write("- **JSON:** Structured format, good for web applications\n")
        f.write("- **Schema:** Metadata and statistics about each dataset\n\n")
        
        f.write("## Usage Examples\n\n")
        f.write("```python\n")
        f.write("import polars as pl\n\n")
        f.write("# Load life table data\n")
        f.write("life_table = pl.read_parquet('life_table.parquet')\n\n")
        f.write("# Load policy data\n")
        f.write("policies = pl.read_parquet('policies.parquet')\n\n")
        f.write("# Perform analysis\n")
        f.write("avg_premium = policies.select(pl.col('Annual_Premium').mean())\n")
        f.write("```\n")
    
    print(f"📄 Summary report created: {report_path}")


def main():
    """Main function to run the conversion."""
    # First create the sample data if it doesn't exist
    excel_path = Path("sample_data/actuarial_data.xlsx")
    
    if not excel_path.exists():
        print("📊 Creating sample actuarial data...")
        # Import and run the data creation script
        import sys
        sys.path.append("examples")
        from create_actuarial_data import create_actuarial_excel_file
        create_actuarial_excel_file()
    
    # Convert to Polars format
    convert_excel_to_polars(
        excel_path="sample_data/actuarial_data.xlsx",
        output_dir="output"
    )


if __name__ == "__main__":
    main()
