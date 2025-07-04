#!/usr/bin/env python3
"""
Quick viewer to show the contents of the actuarial Excel file.
"""

from pathlib import Path

import polars as pl
from openpyxl import load_workbook


def view_actuarial_data():
    """Display contents of the actuarial Excel file."""
    excel_path = "sample_data/actuarial_data.xlsx"
    
    if not Path(excel_path).exists():
        print(f"❌ File not found: {excel_path}")
        print("Please run: python3 examples/create_actuarial_data.py first")
        return
    
    print("📊 ACTUARIAL DATA EXCEL FILE CONTENTS")
    print("=" * 50)
    
    # Get sheet names using openpyxl
    try:
        workbook = load_workbook(filename=excel_path, read_only=True)
        sheets = workbook.sheetnames
        workbook.close()
    except Exception as e:
        print(f"❌ Error reading file: {e}")
        return
    
    print(f"📋 Found {len(sheets)} sheets: {', '.join(sheets)}")
    
    # Display each sheet
    for i, sheet_name in enumerate(sheets, 1):
        print(f"\n{i}. 📑 SHEET: {sheet_name}")
        print("-" * 40)
        
        try:
            # Read sheet with polars
            df = pl.read_excel(
                source=excel_path,
                sheet_name=sheet_name,
                has_header=True
            )
            
            print(f"   📊 Shape: {df.shape} (rows × columns)")
            print(f"   📝 Columns: {', '.join(df.columns)}")
            
            # Show data types
            print("   🏗️  Data Types:")
            for col, dtype in zip(df.columns, df.dtypes):
                print(f"      {col}: {dtype}")
            
            # Show sample data
            print("   📄 Sample Data (first 3 rows):")
            print(str(df.head(3)))
            
            # Show basic statistics for numeric columns
            numeric_cols = [col for col, dtype in zip(df.columns, df.dtypes) 
                           if dtype in [pl.Int64, pl.Int32, pl.Float64, pl.Float32]]
            
            if numeric_cols:
                print("   📈 Basic Statistics (numeric columns):")
                try:
                    stats_df = df.select(numeric_cols).describe()
                    print(str(stats_df))
                except Exception as e:
                    print(f"      Error calculating statistics: {e}")
            
        except Exception as e:
            print(f"   ❌ Error reading sheet {sheet_name}: {e}")
    
    print(f"\n✨ File location: {Path(excel_path).absolute()}")
    print("💡 This file contains realistic actuarial data for:")
    print("   • Life tables with mortality rates")
    print("   • Insurance policy portfolios") 
    print("   • Claims data")
    print("   • Reserve calculations")


if __name__ == "__main__":
    view_actuarial_data()
